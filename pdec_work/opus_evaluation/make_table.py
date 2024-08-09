import openpyxl
import os
import numpy as np
import sys

def _read_txt_strip_(url):
    file = open(url, 'r', encoding='utf-8')
    lines = file.readlines()
    file.close()
    return [line.strip() for line in lines]
    
def make_bar(sheet, start_position, metric, en_results, zero_results):
    sheet.cell(row=1, column=(start_position+1)).value = metric
    sheet.cell(row=2, column=start_position).value = "en2m"
    sheet.cell(row=2, column=(start_position+1)).value = "m2en"
    sheet.cell(row=2, column=(start_position+2)).value = "zero"
    sheet.cell(row=3, column=start_position).value = round(np.sum(en_results[0, :]) / np.count_nonzero(en_results[0, :]), 2)
    sheet.cell(row=3, column=(start_position+1)).value = round(np.sum(en_results[1, :]) / np.count_nonzero(en_results[1, :]), 2)
    sheet.cell(row=3, column=(start_position+2)).value = round(np.sum(zero_results) / np.count_nonzero(zero_results), 2)

def extract_results(model_id, dir_path, metric, language_dict, zero_dict):
    keys_mapping = {
        "sacrebleu": "score",
        "chrf": "score",
        "bertscore": "F: ",
        "comet": "Score",
    }
    data = _read_txt_strip_(os.path.join(dir_path, str(model_id), f"{model_id}.{metric}"))
    en_results = np.zeros((2, len(language_dict)))
    zero_results = np.zeros((len(zero_dict), len(zero_dict)))

    src, tgt = None, None
    for row in data:
        if "-" in row and len(row) < 8:
            tmp = row.split("-")
            src, tgt = tmp[0], tmp[1]
        if f"{keys_mapping[metric]}" in row:
            if metric == "bertscore":
                score = float(row.split(":")[3])
                score = round(score, 2)
            elif metric == "sacrebleu" or metric == "chrf":
                score = float(row.split(":")[2].split(",")[0])
                score = round(score, 2)
            elif metric == "comet":
                score = float(row.split(":")[1])
                score = round(score, 2)
            
            if src == "en":
                en_results[0][language_dict[tgt] - 1] = score
            elif tgt == "en":
                en_results[1][language_dict[src] - 1] = score
            elif src != "en" and tgt != "en":
                zero_results[zero_dict[src] - 1][zero_dict[tgt] - 1] = score
    return en_results, zero_results


def make_detailed_table(sheet, start_row, metric, language_dict, zero_dict, en_results, zero_results):
    language_list = list(language_dict.keys())

    sheet.cell(row=start_row, column=1).value = metric
    sheet.cell(row=(start_row + 1), column=1).value = "en2m"
    sheet.cell(row=(start_row + 4), column=1).value = "m2en"
    for lang in language_list:
        idx_lang = language_dict[lang]
        sheet.cell(row=(start_row + 2), column=idx_lang).value = lang
        sheet.cell(row=(start_row + 3), column=idx_lang).value = en_results[0][idx_lang - 1]
        sheet.cell(row=(start_row + 5), column=idx_lang).value = lang
        sheet.cell(row=(start_row + 6), column=idx_lang).value = en_results[1][idx_lang - 1]
    
    sheet.cell(row=(start_row + 7), column=1).value = "zero"
    zero_list = list(zero_dict.keys())

    for src in zero_list:
        for tgt in zero_list:
            if src == tgt: continue
            src_idx, tgt_idx = zero_dict[src], zero_dict[tgt]
            sheet.cell(row=(start_row + 7) + src_idx, column=tgt_idx).value = zero_results[src_idx - 1][tgt_idx - 1]

def mk_table(id, method_name, work_path, metric_list, language_dict, bertscore_dict, comet_dict, zero_dict):
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(index=0, title="sheet1")
    # bleu
    count_column = 6
    iteration_list = [language_dict, language_dict, bertscore_dict, comet_dict]
    for i in range(len(metric_list)):
        metric = metric_list[i]
        en_results, zero_results = extract_results(id, os.path.join(work_path, method_name), metric_list[i], iteration_list[i], zero_dict)
        make_bar(sheet, (i * 3 + 1), metric, en_results, zero_results)
        make_detailed_table(sheet, count_column, metric, iteration_list[i], zero_dict, en_results, zero_results)
        # 3 is span, (7 + len(zero_dict)) is the length of table
        count_column = count_column + 10 + len(zero_dict)

    save_dir = os.path.join(work_path, "excel", method_name)
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    wb.save(os.path.join(save_dir, str(id) + ".xlsx"))

# all excluding English
language_sequence = ["af", "am", "ar", "as", "az", "be", "bg",
                     "bn", "br", "bs", "ca", "cs", "cy", "da", 
                     "de", "el", "zu", "eo", "es", "et", "eu", 
                     "fa", "fi", "fr", "fy", "ga", "gd", "gl", 
                     "gu", "ha", "he", "hi", "hr", "hu", "id", 
                     "ig", "is", "it", "ja", "ka", "kk", "km", 
                     "kn", "ko", "ku", "ky", "li", "lt", "lv", 
                     "mg", "mk", "ml", "mr", "ms", "mt", "my", 
                     "nb", "ne", "nl", "nn", "no", "oc", "or", 
                     "pa", "pl", "ps", "pt", "ro", "ru", "rw", 
                     "se", "sh", "si", "sk", "sl", "sq", "sr", 
                     "sv", "ta", "te", "tg", "th", "tk", "tr", 
                     "tt", "ug", "uk", "ur", "uz", "vi", "wa", 
                     "xh", "yi", "zh",]
bertscore_sequence = ["af", "ar", "az", "be", "bg", "bn", "br", "bs", "ca", "cs", "cy", "da",
                    "de", "el", "es", "et", "eu", "fi", "fr", "ga", "gl", "gu", "he", "nn",
                    "hi", "hr", "hu", "id", "is", "it", "ja", "ka", "kk", "kn", "ko", "lt",
                    "lv", "mg", "mk", "ml", "mr", "ms", "my", "ne", "nl", "oc", "pa", "pl",
                    "pt", "ro","ru", "sh", "sk", "sl", "sq", "sr", "sv", "ta", "te", "tg", 
                    "tr", "tt", "uk", "ur", "uz", "vi", "zh", "no", "nb", ]
comet_sequence = ["af", "am", "ar", "as", "az", "be", "bg", "bn", "br",
                     "bs", "ca", "cs", "cy", "da", "de", "el", "nb", "eo",
                     "es", "et", "eu", "fa", "fi", "fr", "fy", "ga", "gd",
                     "gl", "gu", "ha", "he", "hi", "hr", "hu", "id", "is",
                     "it", "ja", "ka", "kk", "km", "kn", "ko", "ku", "ky",
                     "lt", "lv", "mg", "mk", "ml", "mr", "ms", "my", "ne",
                     "nl", "no", "or", "pa", "pl", "ps", "pt", "ro", "ru",
                     "si", "sk", "sl", "sq", "sr", "sv", "ta", "te", "th", 
                     "tr", "ug", "uk", "ur", "uz", "vi", "xh", "yi", "zh", "nn",]
zero_sequence = ["de", "nl", "zh", "ar", "ru", "fr"]

language_dict = {value: (idx + 1) for idx, value in enumerate(language_sequence)}
bertscore_dict = {value: (idx + 1) for idx, value in enumerate(bertscore_sequence)}
comet_dict = {value: (idx + 1) for idx, value in enumerate(comet_sequence)}
zero_dict = {value: (idx + 1) for idx, value in enumerate(zero_sequence)}



metric_list = ["sacrebleu", "chrf", "bertscore", "comet"]

method_name = sys.argv[1]
id = sys.argv[2]
root_path = sys.argv[3]

work_path = os.path.join(root_path, "pdec_work", "results")

mk_table(id, method_name, work_path, metric_list, language_dict, bertscore_dict, comet_dict, zero_dict)