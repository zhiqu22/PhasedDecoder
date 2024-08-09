import openpyxl
import os
import numpy as np
import sys

def _read_txt_strip_(url):
    file = open(url, 'r', encoding='utf-8')
    lines = file.readlines()
    file.close()
    return [line.strip() for line in lines]

def extract_results_from_file(model_id, dir_path, metric, language_num, language_dict):
    keys_mapping = {
        "sacrebleu": "score",
        "chrf": "score",
        "bertscore": "F: ",
        "comet": "Score",
    }
    data = _read_txt_strip_(os.path.join(dir_path, str(model_id), f"{model_id}.{metric}"))
    score_list = np.zeros((language_num, language_num))
    idx_i, idx_j = '', ''
    for row in data:
        if "-" in row and len(row) < 8:
            tmp = row.split("-")
            idx_i, idx_j = language_dict[tmp[0]] - 1, language_dict[tmp[1]] - 1
        if f"{keys_mapping[metric]}" in row:
            if metric == "bertscore":
                score = float(row.split(":")[3])
                score = round(score, 2)
            elif metric == "sacrebleu" or metric == "chrf":
                score = float(row.split(":")[2].split(",")[0])
                score = round(score, 2)
            elif metric == "comet":
                score = float(row.split(":")[1])
                score = round(score, 2)
            score_list[idx_i][idx_j] = score
    return score_list
    
def make_bar(sheet, start_position, metric, scores):
    sheet.cell(row=1, column=(start_position+1)).value = metric
    sheet.cell(row=2, column=start_position).value = "en2m"
    sheet.cell(row=2, column=(start_position+1)).value = "m2en"
    sheet.cell(row=2, column=(start_position+2)).value = "zero"
    sheet.cell(row=3, column=start_position).value = round(np.sum(scores[0, :]) / np.count_nonzero(scores[0, :]), 2)
    sheet.cell(row=3, column=(start_position+1)).value = round(np.sum(scores[:, 0]) / np.count_nonzero(scores[:, 0]), 2)
    sheet.cell(row=3, column=(start_position+2)).value = round(np.sum(scores[1:, 1:]) / np.count_nonzero(scores[1:, 1:]), 2)

def make_detailed_table(sheet, start_row, metric, scores, language_sequence):
    sheet.cell(row=start_row, column=1).value = metric
    sheet.cell(row=(start_row + 1), column=1).value = "src"
    sheet.cell(row=start_row, column=2).value = "tgt"
    language_num = len(language_sequence)
    for i in range(language_num):
        sheet.cell(row=(start_row + 1), column=(2 + i)).value = language_sequence[i]
        sheet.cell(row=(start_row + 2 + i), column=1).value = language_sequence[i]
        for j in range(language_num):
            if scores[i][j] == 0: continue
            sheet.cell(row=(start_row + 2) + i, column=(2 + j)).value = scores[i][j]

def mk_table(id, method_name, work_path, metric_list, language_sequence, language_dict):
    language_num = len(language_sequence)
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(index=0, title="sheet1")

    for i in range(len(metric_list)):
        metric = metric_list[i]
        scores = extract_results_from_file(id, os.path.join(work_path, method_name), metric, language_num, language_dict)
        make_bar(sheet, (i * 3 + 1), metric, scores)
        make_detailed_table(sheet, (i * (len(language_sequence) + 3) + 6), metric, scores, language_sequence)

    save_dir = os.path.join(work_path, "excel", method_name)
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    
    wb.save(os.path.join(save_dir, str(id) + ".xlsx"))

language_sequence = ["en", "ar", "he", "ru", "ko", "it", "ja", "zh", "es", "nl", "vi", "tr", "fr", "pl", "ro", "fa", "hr", "cs", "de"]
language_dict = {'en': 1, 'ar': 2, 'he': 3, 'ru': 4, 'ko': 5, 'it': 6, 'ja': 7,
                         'zh': 8, 'es': 9, 'nl': 10, "vi": 11, "tr": 12, "fr":13,
                         "pl": 14, "ro":15, "fa":16, "hr":17, "cs":18, "de": 19}
metric_list = ["sacrebleu", "chrf", "bertscore", "comet"]

method_name = sys.argv[1]
id = sys.argv[2]
root_path = sys.argv[3]

work_path = os.path.join(root_path, "pdec_work", "results")
mk_table(id, method_name, work_path, metric_list, language_sequence, language_dict)